import sys
from pathlib import Path

from openpyxl import load_workbook


if len(sys.argv) != 2:
    raise SystemExit("Usage: python verify_zt_left_excel.py <xlsx>")

path = Path(sys.argv[1])
workbook = load_workbook(path, read_only=True, data_only=True)
found = []
for ws in workbook.worksheets:
    if ws.title == "概要":
        continue
    headers = {ws.cell(5, col).value: col for col in range(2, ws.max_column + 1)}
    required = ["見出し", "左側接続先候補", "読取状態", "確認状態"]
    if not all(key in headers for key in required):
        continue
    for row in range(6, ws.max_row + 1):
        header = str(ws.cell(row, headers["見出し"]).value or "")
        left = str(ws.cell(row, headers["左側接続先候補"]).value or "")
        state = str(ws.cell(row, headers["読取状態"]).value or "")
        confirmation = str(ws.cell(row, headers["確認状態"]).value or "")
        if header.upper().replace(" ", "").startswith("ZT") and "LEFT-" in left.upper():
            found.append((ws.title, row, header, left, state, confirmation))

print(f"ZT LEFT candidates: {len(found)}")
for item in found:
    print(" | ".join(str(value) for value in item))
if not found:
    raise AssertionError("No ZT LEFT row was found in this test Excel; PDF extraction coverage must be improved before claiming the real-PDF test.")
if any(item[4] == "対象外" or item[5] == "対象外" for item in found):
    raise AssertionError("A ZT LEFT row was marked excluded.")
print("ZT LEFT Excel verification: PASSED")
