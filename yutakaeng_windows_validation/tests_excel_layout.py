import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent / "src"))
from pipeline import WireRow, compose_mark_text, write_excel


def check(actual, expected, label):
    if actual != expected:
        raise AssertionError(f"{label}: expected={expected!r} actual={actual!r}")
    print(f"OK  {label}")


with tempfile.TemporaryDirectory() as temp_dir:
    output_dir = Path(temp_dir)
    rows = [
        WireRow(1, "P1-F1", "YF", "マルチコネクター", "1", "FL1A11", "L-A", "R-Y2", ["Y2"], "未確認", "", "", "", "5A"),
        WireRow(1, "P1-F1", "YF", "マルチコネクター", "2", "", "L-B", "R-Y2", ["Y2"], "警告あり", "", "主文字未読取", "", "5A"),
        WireRow(1, "P1-F2", "TF", "マルチコネクター", "1", "TB2F1A1", "", "", [], "未確認", "", "", "", "5A"),
        WireRow(2, "P2-F2", "ZF", "コネクター", "1", "BTB2C6", "", "R-Y2", ["Y2"], "未確認", "", "", "", "6A"),
        WireRow(2, "P2-F3", "IF", "コネクター", "1", "IF1", "L-UNKNOWN", "R-UNKNOWN", [], "警告あり", "", "", "", "6A"),
        WireRow(1, "P1-F4", "ZT1(14)", "端子台", "1", "ZTLINE1", "RIGHT-Y5", "LEFT-Y5", ["Y5"], "未確認", "", "", "", "5A"),
    ]
    output = write_excel(rows, Path("sample.pdf"), "24JNF09501", output_dir, lambda *_: None, page_panels={1: "5A", 2: "6A"})
    workbook = load_workbook(output, data_only=True)
    worksheet = workbook["2"]
    check("Y2" in workbook.sheetnames, False, "wire color is not used in a sheet name")
    check("2" in workbook.sheetnames, True, "wire size alone is used in a sheet name")
    check(worksheet.cell(1, 1).value, "24JNF09501", "order number is the first cell")
    check(worksheet.cell(2, 1).value, "5A", "first page panel number is the second cell")
    expected_headers = ["マーク主文字", "マーク個数", "読取状態", "確認状態", "L側接続先", "R側接続先", "線コード", "見出し", "種別", "PDFページ", "枠ID", "行番号", "警告・除外理由"]
    check([cell.value for cell in worksheet[5]], expected_headers, "hot-marker column order")
    check([worksheet.cell(6, column).value for column in range(1, 4)], ["5A", None, None], "first page panel section row")
    check([worksheet.cell(7, column).value for column in range(1, 4)], ["YF", None, None], "standalone header row")
    check([worksheet.cell(8, column).value for column in range(1, 8)], ["FL1A11", 2, "読取済み", "確認不要", "L-A", "R-Y2", "Y2"], "main row below header")
    check([worksheet.cell(9, column).value for column in range(1, 4)], [None, 0, "警告あり"], "unread main text receives zero marks")
    check(compose_mark_text("3A", "T1(2)", "BTBT2BP"), "BTBT2BP", "terminal display separated from mark cell")
    check(worksheet.cell(10, 1).value, "盤番号: 6A", "second page panel section row")
    size_missing = workbook["線サイズ未記載"]
    check(size_missing.cell(1, 1).value, "24JNF09501", "missing size sheet keeps order number")
    check([size_missing.cell(6, column).value for column in range(1, 4)], ["5A", None, None], "missing size panel section row")
    check([size_missing.cell(7, column).value for column in range(1, 4)], ["TF", None, None], "missing size standalone header row")
    check([size_missing.cell(8, column).value for column in range(1, 5)], ["TB2F1A1", 2, "警告あり", "警告確認"], "missing size row remains editable warning")
    check(size_missing.cell(8, 13).value, "線サイズ未記載", "missing size reason is explicit")
    size_unclear = workbook["線サイズ判別不明"]
    check(size_unclear.cell(1, 1).value, "24JNF09501", "unclear size sheet keeps order number")
    check([size_unclear.cell(6, column).value for column in range(1, 4)], ["6A", None, None], "unclear size panel section row")
    check([size_unclear.cell(7, column).value for column in range(1, 4)], ["IF", None, None], "unclear size standalone header row")
    check([size_unclear.cell(8, column).value for column in range(1, 5)], ["IF1", 2, "警告あり", "警告確認"], "unclear size row remains warning")
    check(size_unclear.cell(8, 13).value, "線サイズ判別不明", "unclear size reason is explicit")
    check(worksheet.freeze_panes, "A6", "top layout is frozen below headers")
    zt_sheet = workbook["ZTブロック"]
    check("Y5" in workbook.sheetnames, False, "ZT-only wire is not mixed into normal size sheet")
    check(zt_sheet.cell(1, 1).value, "24JNF09501", "ZT sheet keeps order number")
    check([zt_sheet.cell(6, column).value for column in range(1, 4)], ["5A", None, None], "ZT sheet keeps page panel section")
    check([zt_sheet.cell(7, column).value for column in range(1, 4)], ["ZT1", None, None], "ZT header is standalone")
    check([zt_sheet.cell(8, column).value for column in range(1, 8)], ["ZTLINE1", 2, "読取済み", "確認不要", "RIGHT-Y5", "LEFT-Y5", "Y5"], "RIGHT and LEFT ZT row is written to dedicated sheet")

    x_output = write_excel(rows, Path("sample_x.pdf"), "要確認", output_dir, lambda *_: None, x_only=True, page_panels={1: "5A", 2: "6A"})
    x_workbook = load_workbook(x_output, data_only=True)
    x_sheet = x_workbook["2"]
    check(x_sheet.cell(8, 2).value, 2, "X-block mode marks readable rows twice")
    check(x_sheet.cell(9, 2).value, 2, "X-block mode marks unread rows twice")

print("ALL EXCEL LAYOUT TESTS PASSED")
