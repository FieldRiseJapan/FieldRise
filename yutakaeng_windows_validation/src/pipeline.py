from __future__ import annotations

import os
import re
import shutil
import tempfile
import time
from collections import defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable

import cv2
import fitz  # PyMuPDF
import numpy as np
import pytesseract
from rapidocr import RapidOCR
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

LogFn = Callable[[str, str], None]
ProgressFn = Callable[[int, int], None]

NAVY = "102A43"
LIGHT = "EAF4FF"
WARN = "FFCCBC"
EXCLUDED = "E8F5E9"
SECTION = "FFF3E0"
THIN = Side(style="thin", color="D6E3F0")
# 数字を含まない主文字は、実PDFで原図照合済みの識別子だけを候補化する。
# これにより従来OCRで発生したRTRP等の英字だけの無意味候補をExcelへ出さない。
ALPHA_ONLY_MAIN_ALLOWLIST = {"BTBP", "BTBN", "SBP", "SBN"}
MAIN_OCR_ENGINE: RapidOCR | None = None


@dataclass
class WireRow:
    page: int
    frame_id: str
    header: str
    kind: str
    row_no: str
    main_text: str
    left_reference: str
    right_reference: str
    wire_codes: list[str]
    state: str
    exclusion_reason: str
    warning: str
    source_crop: str
    panel_no: str = ""


def app_root() -> Path:
    """持ち運び版では実行ファイル横、通常実行ではプロジェクト直下を返す。"""
    import sys
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parents[1]


def configure_tesseract() -> None:
    if os.name != "nt":
        return
    bundled = app_root() / "Tesseract-OCR" / "tesseract.exe"
    if bundled.exists():
        pytesseract.pytesseract.tesseract_cmd = str(bundled)
        os.environ.setdefault("TESSDATA_PREFIX", str(app_root() / "Tesseract-OCR" / "tessdata"))


def clean_text(value: str) -> str:
    value = (value or "").replace("\n", " ").replace("\r", " ")
    value = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", value)
    return re.sub(r"\s+", " ", value).strip()


def ocr(image: np.ndarray, psm: int = 7, whitelist: str = "", scale: float = 3.0) -> str:
    if image.size == 0:
        return ""
    enlarged = cv2.resize(image, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    config = f"--oem 1 --psm {psm}"
    if whitelist:
        config += f" -c tessedit_char_whitelist={whitelist}"
    text = pytesseract.image_to_string(enlarged, config=config, lang="eng")
    return clean_text(text)


def normalize_header(text: str) -> str:
    normalized = clean_text(text).upper().replace("=", "-").replace(" ", "")
    # 小型ZT枠はZ7T1のようにZとTの間へノイズが入ることがある。
    # 見出し用途ではZT1として正規化するが、行内の文字は変更しない。
    zt = re.search(r"Z.?T[1I](?:\(\d+\))?", normalized)
    if zt:
        return re.sub(r"^Z.?T", "ZT", zt.group(0)).replace("I", "1")
    return normalized


def classify_header(header: str) -> str:
    h = normalize_header(header)
    base = re.split(r"[-(]", h)[0]
    if base in {"YF", "OF", "SF", "BSF", "RJF", "TF", "PF", "RPF"}:
        return "マルチコネクター"
    if base in {"LF", "RF", "FF", "ZF", "WF", "IF", "BTF", "BZF", "BWF"}:
        return "コネクター"
    if re.match(r"^(ZT|T|X|FT|RT|DT)\d", base):
        return "端子台"
    return "未分類"


def wire_size_status(*values: str) -> str:
    """線サイズの状態を、未記載と記載あり・判別不明に分ける。

    文字がまったくない場合は未記載、文字や数字はあるが安全な色+サイズ構文に
    一致しない場合は判別不明とする。推測によるサイズ補正は行わない。
    """
    texts = [clean_text(value).upper().replace(" ", "") for value in values if clean_text(value)]
    meaningful = [text for text in texts if text not in {"UNCLEAR", "未判読"}]
    if not meaningful:
        return "未記載"
    if extract_wire_codes(*meaningful):
        return "判別済み"
    return "判別不明"


def extract_wire_codes(*values: str) -> list[str]:
    """左右接続先に付く線コードだけを抽出する。主文字・接続先番号はシート名に使わない。"""
    codes: list[str] = []
    # RIGHT-Y5、LEFT-Y5、R-Y2、D-Y0.5、T1:1-Y2等から、色+サイズだけを読む。
    # A01等のコネクター端子参照を電線サイズと誤認しないため、既知の色文字だけを許可する。
    # 現時点で確認済みの色表記。OCRの0/O混同によるBO2・YO2等は未分類へ残す。
    color = r"(?:GY|Y|G|K|B|W|R)"
    color_size = re.compile(rf"(?<![A-Z0-9])((?:{color}[0-9]+(?:[.][0-9]+)?|[0-9]+(?:[.][0-9]+)?{color}))(?![A-Z0-9])")
    approved_sizes = {"0.3", "0.5", "0.75", "1.25", "1.5", "2", "3", "3.5", "5", "5.5", "8", "14", "22", "38"}
    for value in values:
        text = clean_text(value).upper()
        for code in color_size.findall(text):
            color_first = re.fullmatch(r"[A-Z]{1,2}([0-9]+(?:[.][0-9]+)?)", code)
            size_first = re.fullmatch(r"([0-9]+(?:[.][0-9]+)?)[A-Z]{1,2}", code)
            size = (color_first or size_first).group(1)
            # 未登録・不自然なOCR値は未分類に残し、誤ったサイズシートを作らない。
            if size not in approved_sizes:
                continue
            if code not in codes:
                codes.append(code)
    return codes


def is_terminal(header: str) -> bool:
    return classify_header(header) == "端子台"


def is_x_block(header: str) -> bool:
    """X1、X2など、Xで始まる端子台ブロックだけを判定する。"""
    normalized = normalize_header(header)
    base = re.split(r"[-(]", normalized)[0]
    return bool(re.fullmatch(r"X\d+", base))


def is_if_block(header: str) -> bool:
    """盤内線モードで対象にするIFブロックを判定する。"""
    normalized = normalize_header(header)
    base = re.split(r"[-(]", normalized)[0]
    return base == "IF"


def is_internal_wire_reference(header: str, left: str, right: str) -> bool:
    """端子台側のI表記を盤内線として判定する。単独のIを英数字ノイズとして拡張しない。"""
    if not is_terminal(header):
        return False
    refs = f"{left}/{right}".upper()
    return bool(re.search(r"(^|[=/])\s*I(?=$|[-=/:.])", refs))


def exclusion_reason(header: str, left: str, right: str, keep_internal: bool = False) -> str:
    h = normalize_header(header)
    if h.startswith("DT"):
        return "対象外: 扉付きDT系端子台"
    refs = f"{left}/{right}".upper()
    if re.search(r"(^|[=/\s])D-", refs):
        return "対象外: 線コードがD-で始まる"
    # ZTブロックのLEFT/RIGHTは盤間配線であり、I除外より優先して対象に残す。
    if h.startswith("ZT") and re.search(r"\b(?:LEFT|RIGHT)-", refs):
        return ""
    if not keep_internal and is_internal_wire_reference(header, left, right):
        return "対象外: 端子台接続先のIは盤内行き"
    return ""


def render_page(page: fitz.Page, dpi: int = 300) -> np.ndarray:
    pix = page.get_pixmap(matrix=fitz.Matrix(dpi / 72, dpi / 72), colorspace=fitz.csGRAY, alpha=False)
    return np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width)


def overlap(a: tuple[int, int, int, int], b: tuple[int, int, int, int]) -> float:
    ax, ay, aw, ah = a
    bx, by, bw, bh = b
    x1, y1 = max(ax, bx), max(ay, by)
    x2, y2 = min(ax + aw, bx + bw), min(ay + ah, by + bh)
    inter = max(0, x2 - x1) * max(0, y2 - y1)
    union = aw * ah + bw * bh - inter
    return inter / union if union else 0.0


def detect_frames(gray: np.ndarray) -> list[tuple[int, int, int, int]]:
    h, w = gray.shape
    binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 41, 9)
    horizontal = cv2.morphologyEx(binary, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_RECT, (max(35, w // 65), 1)))
    vertical = cv2.morphologyEx(binary, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(35, h // 65))))
    grid = cv2.dilate(cv2.bitwise_or(horizontal, vertical), cv2.getStructuringElement(cv2.MORPH_RECT, (7, 7)), 1)
    mx, my = int(w * 0.025), int(h * 0.025)
    grid[:my] = grid[h - my:] = 0
    grid[:, :mx] = grid[:, w - mx:] = 0
    contours, _ = cv2.findContours(grid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    raw = []
    for c in contours:
        x, y, fw, fh = cv2.boundingRect(c)
        if fw < w * 0.026 or fh < h * 0.038 or fw * fh < w * h * 0.0015:
            continue
        if fw > w * 0.50 or fh > h * 0.55 or (y > h * 0.83 and x > w * 0.57):
            continue
        if 0.06 <= fw / fh <= 5.2:
            raw.append((x, y, fw, fh))
    result: list[tuple[int, int, int, int]] = []
    for frame in sorted(raw, key=lambda item: item[2] * item[3]):
        if all(overlap(frame, previous) < 0.72 for previous in result):
            result.append(frame)
    # ZTブロックは高さが小さく、一般枠の最小高さ条件だけでは見逃すことがある。
    # 見出しZT1等を局所OCRで探し、その直下を専用枠候補として追加する。
    data = pytesseract.image_to_data(gray, config="--oem 1 --psm 11", output_type=pytesseract.Output.DICT)
    zt_headers = []
    for index, raw in enumerate(data["text"]):
        text = clean_text(raw).upper().replace(" ", "")
        if re.fullmatch(r"ZT[1I]?(?:\(\d+\))?", text):
            zt_headers.append((data["left"][index], data["top"][index], data["width"][index], data["height"][index]))
    zt_headers.sort(key=lambda item: (item[0], item[1]))
    for header_x, header_y, header_w, header_h in zt_headers:
        below = [item[1] for item in zt_headers if item[0] == header_x and item[1] > header_y]
        next_y = min(below) if below else header_y + 420
        body_y = header_y + header_h + 6
        body_h = min(380, max(80, next_y - body_y - 8))
        candidate = (max(0, header_x - 2), body_y, min(w - header_x + 2, max(260, header_w + 220)), body_h)
        if all(overlap(candidate, previous) < 0.72 for previous in result):
            result.append(candidate)
    return sorted(result, key=lambda item: (item[1], item[0]))


def crop(gray: np.ndarray, x: int, y: int, w: int, h: int) -> np.ndarray:
    height, width = gray.shape
    return gray[max(0, y):min(height, y + h), max(0, x):min(width, x + w)]


def remove_grid_lines(image: np.ndarray) -> np.ndarray:
    """文字を残しつつ、セル境界の横罫線と端の縦罫線だけを除去する。"""
    if image.size == 0:
        return image
    binary = cv2.threshold(image, 200, 255, cv2.THRESH_BINARY_INV)[1]
    # 文字の横棒を消さないよう、セル幅の大半を横切る線だけを除去する。
    horizontal = cv2.morphologyEx(binary, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_RECT, (max(24, int(image.shape[1] * .75)), 1)))
    cleaned = cv2.subtract(binary, horizontal)
    # 端の縦罫線を白にする。中央のI・1・Lなどの文字は絶対に消さない。
    margin = max(2, min(10, image.shape[1] // 18))
    cleaned[:, :margin] = 0
    cleaned[:, max(margin, image.shape[1] - margin):] = 0
    return cv2.bitwise_not(cleaned)


def main_ocr_engine() -> RapidOCR:
    """主文字専用のローカルONNX認識器を一度だけ初期化する。"""
    global MAIN_OCR_ENGINE
    if MAIN_OCR_ENGINE is None:
        MAIN_OCR_ENGINE = RapidOCR()
    return MAIN_OCR_ENGINE


def normalize_main_candidate(value: str) -> str:
    """主文字に使える英数字・業務上の記号だけを残し、空白を除く。"""
    return clean_text(value).upper().replace(" ", "")


def main_candidate_is_safe(value: str, confidence: float) -> bool:
    """意味のないOCR文字列をExcelへ出さないための保守的な品質ゲート。"""
    text = normalize_main_candidate(value)
    if confidence < 0.85 or not (2 <= len(text) <= 18):
        return False
    if not re.fullmatch(r"[A-Z0-9\-_/.:+]+", text):
        return False
    alnum = sum(char.isalnum() for char in text)
    # 図面の主文字は少なくとも英字を含む識別子として扱う。単独の端子番号や罫線ノイズは拒否する。
    if alnum < 2 or not re.search(r"[A-Z]", text):
        return False
    # 英字だけの候補は、実証済みの識別子以外を安全側で未読取にする。
    if not re.search(r"\d", text) and text not in ALPHA_ONLY_MAIN_ALLOWLIST:
        return False
    if re.fullmatch(r"([A-Z0-9])\1{2,}", text):
        return False
    # 主文字セル端の罫線・端子番号が混じった実測パターンは、もっともらしく見えても確定しない。
    if re.fullmatch(r"[A-Z]{2}\d{3}S\d{2}L", text):
        return False
    # AC系で先頭Aが欠けた形（C711S11等）は推測復元せず、原図確認へ回す。
    if re.fullmatch(r"[A-Z]\d{3}S\d{2}", text):
        return False
    return True


def rapidocr_main_candidate(image: np.ndarray) -> tuple[str, float]:
    """検出器を使わず、既知の中央セルだけを高速に認識する。"""
    if image.size == 0:
        return "", 0.0
    output = main_ocr_engine()(image, use_det=False, use_cls=False, use_rec=True)
    if not output.txts or not output.scores:
        return "", 0.0
    return normalize_main_candidate(output.txts[0]), float(output.scores[0])


def choose_main_candidate(primary: tuple[str, float], narrow: tuple[str, float]) -> tuple[str, str]:
    """幅の異なる同一セル読取を照合し、一致・安全な末尾補完だけを採用する。"""
    first, first_confidence = primary
    second, second_confidence = narrow
    first_ok = main_candidate_is_safe(first, first_confidence)
    second_ok = main_candidate_is_safe(second, second_confidence)
    if first_ok and second_ok:
        if first == second:
            return first, ""
        # 狭い切出しで末尾だけ欠ける場合に限り、広い切出しの長い一致候補を採用する。
        if first.startswith(second) and len(first) - len(second) <= 2:
            return first, ""
        if second.startswith(first) and len(second) - len(first) <= 2:
            return second, ""
        return "", "主文字候補不一致"
    if first_ok and not second:
        return first, "主文字単独候補（要確認）"
    if second_ok and not first:
        return second, "主文字単独候補（要確認）"
    return "", "主文字未読取"


def read_main_text(gray: np.ndarray, frame: tuple[int, int, int, int], top: int, row_h: int) -> tuple[str, str]:
    """行番号・丸囲み・右端端子番号を避けた主文字セルだけをローカルONNX OCRで読む。"""
    x, _, w, _ = frame
    crop_y = top - 3
    crop_h = row_h + 6
    # 最新実PDFのYF/ZF-B/IFで検証済み。旧方式(.20-.82)は末尾切れや丸囲み混入を起こした。
    broad = crop(gray, x + int(.30 * w), crop_y, max(20, int(.65 * w)), crop_h)
    narrow = crop(gray, x + int(.30 * w), crop_y, max(20, int(.62 * w)), crop_h)
    return choose_main_candidate(rapidocr_main_candidate(broad), rapidocr_main_candidate(narrow))


def normalize_side_candidate(value: str) -> str:
    """図面フォントで生じる中点・ダッシュの表記ゆれを接続先の記号へ戻す。"""
    text = normalize_main_candidate(value)
    text = text.translate(str.maketrans({"·": ":", "—": "-", "–": "-", "−": "-", "一": "-", "×": "X"}))
    text = re.sub(r"[-:.]+$", "", text)
    # RIGHT-Y5上、LEFT-Y5一のように、正しい盤間配線コードの末尾だけへ
    # 罫線・非ASCIIノイズが付く場合は、承認済みの線色・サイズに完全一致する部分だけ残す。
    # 数字が続く値（RIGHT-Y51等）はY5と推測補正しない。
    size = r"(?:0[.]3|0[.]5|0[.]75|1[.]25|1[.]5|3[.]5|5[.]5|2|3|5|8|14|22|38)"
    cross_panel = re.match(rf"^((?:LEFT|RIGHT)-(?:GY|Y|G|K|B|W|R){size})(?:[^A-Z0-9].*)?$", text)
    if cross_panel:
        return cross_panel.group(1)
    return text


def side_candidate_is_safe(value: str, confidence: float) -> bool:
    """接続先として明確に読めた文字列だけを採用する。"""
    if confidence < .85 or not re.fullmatch(r"[A-Z0-9#:/\-.]+", value):
        return False
    # ZT/Xなどの先頭欠落・誤字は推測補正せず、空欄と警告で原図確認へ回す。
    # ただしT1:4のようにコロンまで読めた端子参照は明確なので残す。
    clear_terminal_reference = re.fullmatch(r"(?:ZT|T|X|FT|RT)\d+:[A-Z0-9]+(?:-[A-Z0-9.]+)?", value)
    return not (re.match(r"^(?:7T|F:|[A-Z]1[.:]?[A-Z0-9])", value) and not clear_terminal_reference)


def read_side_reference(gray: np.ndarray, frame: tuple[int, int, int, int], top: int, row_h: int, side: str) -> tuple[str, str]:
    """主文字枠の外側だけを認識し、隣列との混在を防ぐ。"""
    x, _, w, _ = frame
    crop_y = top - 3
    crop_h = row_h + 6
    if side == "left":
        image = crop(gray, x - int(.60 * w), crop_y, max(20, int(.60 * w) - 8), crop_h)
        label = "左側接続先"
    else:
        image = crop(gray, x + w + 5, crop_y, max(20, int(.72 * w)), crop_h)
        label = "右側接続先"
    value, confidence = rapidocr_main_candidate(image)
    value = normalize_side_candidate(value)
    if not value:
        return "", ""
    if not side_candidate_is_safe(value, confidence):
        return "", f"{label}未読取"
    return value, ""


def row_bounds(frame: tuple[int, int, int, int], gray: np.ndarray) -> list[tuple[int, int]]:
    x, y, w, h = frame
    inner = crop(gray, x + int(w * .06), y + 3, int(w * .88), h - 6)
    binary = cv2.threshold(inner, 185, 255, cv2.THRESH_BINARY_INV)[1]
    lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_RECT, (max(20, w // 6), 1)))
    projection = (lines > 0).sum(axis=1)
    points = [index for index, score in enumerate(projection) if score > max(8, w * .25)]
    groups: list[list[int]] = []
    for point in points:
        if not groups or point > groups[-1][-1] + 3:
            groups.append([point])
        else:
            groups[-1].append(point)
    borders = [y + int(sum(group) / len(group)) for group in groups]
    if len(borders) < 3:
        # 枠の高さに合わせた安全な暫定分割。行数は確定しない。
        return [(y + int(h * i / 12), y + int(h * (i + 1) / 12)) for i in range(12)]
    result = []
    for first, second in zip(borders, borders[1:]):
        if second - first >= 12:
            result.append((first + 2, second - 2))
    return result


def normalize_header_ocr_candidate(value: str) -> list[str]:
    """見出し位置で確認済みの限定的な1文字誤読だけを候補化し、推測拡張しない。"""
    text = normalize_main_candidate(value).replace("¥", "X").replace("×", "X")
    text = re.sub(r"^(X\d)\.", r"\1", text)
    text = re.sub(r"^[1I]1\(", "T1(", text)
    candidates = [text]
    known = ("YF", "OF", "SF", "BSF", "RJF", "TF", "PF", "RPF", "LF", "RF", "FF", "ZF", "WF", "IF", "BTF", "BZF", "BWF")
    # 例: YE→YF、SE-A→SF-A、BSE-A→BSF-A。文字数が同じで差が1文字だけの場合のみ。
    tokens = [text] + re.findall(r"[A-Z]{2,4}(?:[-(][A-Z0-9.]+)?", text)
    for token in tokens:
        base = re.split(r"[-(]", token)[0]
        for expected in known:
            if len(base) == len(expected) and sum(a != b for a, b in zip(base, expected)) == 1:
                candidates.append(expected + token[len(base):])
    return candidates


def read_frame_header(gray: np.ndarray, frame: tuple[int, int, int, int]) -> tuple[str, str]:
    """枠直上の短い領域だけから、既知のコネクター・端子台見出しを安全に判定する。"""
    x, y, w, _ = frame
    # 前枠の最終行を含めない狭い見出し帯。最新実PDFでYF/ZF-B/T1/SF-A/BSF-A/IF/X1を確認済み。
    image = crop(gray, x - int(.15 * w), max(0, y - 58), int(1.30 * w), 52)
    detected = main_ocr_engine()(image, use_det=True, use_cls=False, use_rec=True)
    raw_candidates = [normalize_main_candidate(text) for text in (detected.txts or [])]
    direct = rapidocr_main_candidate(image)[0]
    if direct:
        raw_candidates.append(direct)
    variants: list[str] = []
    for candidate in raw_candidates:
        variants.extend(normalize_header_ocr_candidate(candidate))
    for candidate in variants:
        # コロンを含む値はT1:4-B5等の接続先であり、枠見出しには採用しない。
        if ":" in candidate:
            continue
        kind = classify_header(candidate)
        if kind != "未分類":
            return normalize_header(candidate), kind
    return "UNCLEAR", "未分類"


def compose_mark_text(panel_no: str, header: str, main_text: str) -> str:
    """マークセルには主文字だけを置き、見出しはExcelのグループ行へ分離する。

    panel_no と header は出力行のメタデータ・見出し行で保持する。端子台表示を
    主文字へ自動連結すると写真の配置と異なり、マーク番号を二重化するため、ここでは
    推測したプレフィックスを付加しない。
    """
    value = normalize_main_candidate(main_text)
    return "" if value in {"", "UNCLEAR"} else value


def analyze_frame(gray: np.ndarray, page_no: int, frame_no: int, frame: tuple[int, int, int, int], crops_dir: Path, review_dir: Path | None = None, panel_no: str = "", internal_only: bool = False, x_only: bool = False) -> list[WireRow]:
    x, y, w, h = frame
    header, kind = read_frame_header(gray, frame)
    rows: list[WireRow] = []
    for index, (top, bottom) in enumerate(row_bounds(frame, gray), 1):
        row_h = max(12, bottom - top)
        # 左右接続先は主文字枠の外側だけ、主文字は丸囲み・端子番号を除いたセルだけを読む。
        left, left_warning = read_side_reference(gray, frame, top, row_h, "left")
        main, main_warning = read_main_text(gray, frame, top, row_h)
        right, right_warning = read_side_reference(gray, frame, top, row_h, "right")
        # 実データがない空行は出力しない。文字が読めた行だけを確認対象として残す。
        if not any((left, main, right)):
            continue
        # 盤内線モードは端子台側のI線に加えて、IFブロック内の全マークを対象にする。
        if internal_only and not (is_internal_wire_reference(header, left, right) or is_if_block(header)):
            continue
        # XブロックモードはXブロック内の全マークを対象にし、通常モードの除外規則を適用しない。
        if x_only and not is_x_block(header):
            continue
        reason = "" if x_only or (internal_only and is_if_block(header)) else exclusion_reason(header, left, right, keep_internal=internal_only)
        warning = "; ".join(item for item in (main_warning, left_warning, right_warning) if item)
        state = "対象外" if reason else "未確認"
        # 主文字の未読取・候補不一致・単独候補は、空欄であっても必ず利用者確認の警告として残す。
        if warning and not reason:
            state = "警告あり"
        if any(value == "UNCLEAR" for value in (header, main, left, right)):
            warning = (warning + "; " if warning else "") + "未判読候補あり"
            if not reason:
                state = "警告あり"
        crop_path = crops_dir / f"P{page_no}_F{frame_no}_R{index}.png"
        source_crop = crop(gray, x - int(1.65*w), top - 12, int(3.2*w), row_h + 24)
        # ページ端の誤検出枠では診断切出しが空になることがある。解析本体を止めない。
        if source_crop.size:
            cv2.imwrite(str(crop_path), source_crop)
            # 外部支援用には、警告行の画像だけを別フォルダへ保存する。
            # PDF全体・他ページ・正常行は保存対象にしない。
            if review_dir is not None and warning and not reason:
                review_dir.mkdir(parents=True, exist_ok=True)
                review_path = review_dir / f"P{page_no}_F{frame_no}_R{index}.png"
                cv2.imwrite(str(review_path), source_crop)
        else:
            warning = (warning + "; " if warning else "") + "行画像切出し不可"
            if not reason:
                state = "警告あり"
        rows.append(WireRow(
            page_no, f"P{page_no}-F{frame_no}", header, kind, str(index), main, left, right,
            extract_wire_codes(left, right), state, reason, warning, str(crop_path) if source_crop.size else "", panel_no,
        ))
    return rows


def extract_panel_number(gray: np.ndarray) -> str:
    """右下の盤番号だけを保守的に抽出する。読めない場合は空欄のまま推測しない。"""
    h, w = gray.shape
    footer = crop(gray, int(w * .72), int(h * .76), int(w * .28), int(h * .24))
    text = ocr(footer, 11).upper()
    candidates = re.findall(r"(?<![A-Z0-9])\d{1,2}[A-Z](?:\(\d+\))?(?![A-Z0-9])", text)
    return candidates[0] if candidates else ""


def extract_order_number(gray: np.ndarray) -> str:
    h, w = gray.shape
    title = crop(gray, int(w * .70), int(h * .72), int(w * .30), int(h * .28))
    text = ocr(title, 11).upper()
    candidates = re.findall(r"(?<![A-Z0-9])[A-Z0-9]{8,16}(?![A-Z0-9])", text)
    candidates = [item for item in candidates if re.search(r"\d", item) and re.search(r"[A-Z]", item)]
    return candidates[0] if candidates else "要確認"


def desktop_path() -> Path:
    target = Path.home() / "Desktop"
    return target if target.exists() else Path.home()


def safe_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", name or "未分類")[:31]
    result, number = base or "未分類", 2
    while result in used:
        suffix = f"_{number}"
        result = base[:31-len(suffix)] + suffix
        number += 1
    used.add(result)
    return result


def style_table(ws, header_row: int, final_row: int, columns: int, start_column: int = 2) -> None:
    for col in range(start_column, start_column + columns):
        cell = ws.cell(header_row, col)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(header_row + 1, final_row + 1):
        for col in range(start_column, start_column + columns):
            cell = ws.cell(row, col)
            cell.border = Border(bottom=THIN)
            cell.font = Font(name="Yu Gothic", size=11)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)


def write_excel(rows: list[WireRow], pdf_path: Path, order_no: str, output_dir: Path, log: LogFn, internal_only: bool = False, x_only: bool = False) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{order_no}.xlsx" if order_no != "要確認" else f"要確認_{pdf_path.stem}_{timestamp}.xlsx"
    path = output_dir / filename
    wb = Workbook()
    overview = wb.active
    overview.title = "概要"
    overview.sheet_view.showGridLines = False
    overview.column_dimensions["A"].width = 3
    overview.column_dimensions["B"].width = 28
    overview.column_dimensions["C"].width = 92
    overview.merge_cells("B2:C2")
    overview["B2"] = "yutakaeng 検証出力"
    overview["B2"].font = Font(size=18, bold=True, color=NAVY)
    overview["B3"] = "警告行だけを確認・修正してください。警告がない行は確認不要です。"
    overview["B3"].alignment = Alignment(wrap_text=True, vertical="center")
    overview.row_dimensions[3].height = 36
    summary = [
        ("入力PDF", str(pdf_path)), ("オーダー番号候補", order_no), ("候補行数", len(rows)),
        ("対象外行数", sum(item.state == "対象外" for item in rows)),
        ("警告行数", sum(item.state == "警告あり" for item in rows)),
        ("出力場所", str(path)),
        ("安全状態", "警告行の確認・修正が完了するまで最終出力不可"),
    ]
    overview["B5"], overview["C5"] = "項目", "結果"
    for index, (name, value) in enumerate(summary, 6):
        overview.cell(index, 2, name); overview.cell(index, 3, value)
    style_table(overview, 5, 5 + len(summary), 2)
    overview["C11"].fill = PatternFill("solid", fgColor=WARN)

    # 見出しを分類できない枠は、マーカー候補へ混ぜず解析保留シートへ隔離する。
    # 対象外と判定した行だけを通常の確認対象から外す。
    usable = [item for item in rows if item.state != "対象外"]
    groups: dict[str, list[WireRow]] = defaultdict(list)
    for item in usable:
        if item.wire_codes and item.kind != "未分類":
            for code in item.wire_codes:
                groups[code].append(item)
        elif wire_size_status(item.left_reference, item.right_reference) == "未記載":
            # 図面に線サイズがない行は、判別不明行と分けて残す。
            groups["線サイズ未記載"].append(item)
        elif wire_size_status(item.left_reference, item.right_reference) == "判別不明":
            # 記載らしき文字はあるが安全な色+サイズ構文でない行は推測せず分離する。
            groups["線サイズ判別不明"].append(item)
        else:
            # 線サイズは読めても見出しが分類不能な場合は、マーカー候補に混ぜず警告一覧へ残す。
            groups["警告一覧"].append(item)

    used = {"概要"}
    for code in sorted(groups, key=lambda value: (value in {"線サイズ未記載", "線サイズ判別不明", "警告一覧"}, value == "警告一覧", value)):
        ws = wb.create_sheet(safe_sheet_name(code, used))
        ws.sheet_view.showGridLines = False
        # ホットマーカー作業に必要な列を先頭へ固定し、原図照合用の情報は右側へまとめる。
        widths = [20, 13, 13, 13, 30, 30, 13, 20, 16, 14, 18, 10, 30]
        for index, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(index)].width = width
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(widths))
        ws.cell(2, 1, f"電線サイズ・コード: {code}（確認用候補）")
        ws.cell(2, 1).font = Font(name="Yu Gothic", size=15, bold=True, color=NAVY)
        ws.cell(4, 1, f"盤番号: {rows[0].panel_no if rows and rows[0].panel_no else '要確認'}")
        ws.cell(4, 1).font = Font(name="Yu Gothic", size=11, bold=True, color=NAVY)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(widths))
        count_note = "盤内線・Xブロックモードでは対象行をすべて2個で出力します。" if (internal_only or x_only) else "マーク個数は安全に読めた別電線1本につき2個です。"
        ws.cell(3, 1, f"{count_note}確認状態は警告行だけを確認・修正してください。")
        ws.cell(3, 1).alignment = Alignment(wrap_text=True)
        headers = ["マーク主文字", "マーク個数", "読取状態", "確認状態", "L側接続先", "R側接続先", "線コード", "見出し", "種別", "PDFページ", "枠ID", "行番号", "警告・除外理由"]
        for col, value in enumerate(headers, 1):
            ws.cell(5, col, value)
        output_row = 6
        section_rows: list[int] = []
        previous_header = ""
        for item in groups[code]:
            header_value = re.sub(r"\(\d+\)$", "", clean_text(item.header).upper()) or "未分類"
            if header_value != previous_header:
                section_rows.append(output_row)
                ws.cell(output_row, 1, header_value)
                output_row += 1
                previous_header = header_value
            mark_count = 2 if (internal_only or x_only) and item.state != "対象外" else (2 if item.main_text and item.state != "対象外" else 0)
            display_main = compose_mark_text(item.panel_no, item.header, item.main_text) if item.main_text else ""
            size_status = wire_size_status(item.left_reference, item.right_reference)
            size_missing = size_status != "判別済み"
            read_state = "警告あり" if size_missing else ("読取済み" if item.state == "未確認" else item.state)
            confirm_state = "対象外" if item.state == "対象外" else ("警告確認" if (item.state == "警告あり" or size_missing) else "確認不要")
            warning_text = item.warning or item.exclusion_reason
            if size_status == "未記載":
                warning_text = f"{warning_text}; 線サイズ未記載" if warning_text else "線サイズ未記載"
            elif size_status == "判別不明":
                warning_text = f"{warning_text}; 線サイズ判別不明" if warning_text else "線サイズ判別不明"
            values = [display_main, mark_count, read_state, confirm_state, item.left_reference, item.right_reference, "/".join(item.wire_codes), item.header, item.kind, item.page, item.frame_id, item.row_no, warning_text]
            for col, value in enumerate(values, 1):
                ws.cell(output_row, col, value)
            output_row += 1
        end = output_row - 1
        style_table(ws, 5, end, len(headers), start_column=1)
        for section_row in section_rows:
            for col in range(1, len(headers) + 1):
                cell = ws.cell(section_row, col)
                cell.fill = PatternFill("solid", fgColor=SECTION)
                cell.font = Font(name="Yu Gothic", size=12, bold=True, color=NAVY)
                cell.alignment = Alignment(vertical="center")
                cell.border = Border(bottom=THIN)
            ws.row_dimensions[section_row].height = 24
        validation = DataValidation(type="list", formula1='"警告確認,確認済み,要修正,確認不要,対象外"', allow_blank=False)
        ws.add_data_validation(validation); validation.add(f"D6:D{end}")
        ws.conditional_formatting.add(f"D6:D{end}", FormulaRule(formula=['D6="要修正"'], fill=PatternFill("solid", fgColor=WARN)))
        ws.conditional_formatting.add(f"D6:D{end}", FormulaRule(formula=['D6="警告確認"'], fill=PatternFill("solid", fgColor=WARN)))
        ws.conditional_formatting.add(f"D6:D{end}", FormulaRule(formula=['D6="対象外"'], fill=PatternFill("solid", fgColor=EXCLUDED)))
        ws.conditional_formatting.add(f"C6:C{end}", FormulaRule(formula=['C6="警告あり"'], fill=PatternFill("solid", fgColor=WARN)))
        ws.freeze_panes = "A6"; ws.auto_filter.ref = f"A5:M{end}"
    if not groups:
        ws = wb.create_sheet("警告一覧")
        ws["B2"] = "電線サイズを特定できない行です。警告理由を確認してください。"
    wb.save(path)
    log("EXCEL_WRITE_DONE", f"デスクトップへ保存しました: {path.name}")
    return path


def run_pipeline(pdf_path: str | Path, log: LogFn, progress: ProgressFn, internal_only: bool = False, x_only: bool = False) -> Path:
    configure_tesseract()
    pdf_path = Path(pdf_path)
    if not pdf_path.exists() or pdf_path.suffix.lower() != ".pdf":
        raise FileNotFoundError("PDFファイルを選択してください。")
    try:
        pytesseract.get_tesseract_version()
    except Exception as exc:
        raise RuntimeError("Tesseract OCRを起動できません。検証版に同梱されたOCR実行ファイルを確認してください。") from exc

    started = time.monotonic()
    work = Path(tempfile.mkdtemp(prefix="yutakaeng_"))
    crops = work / "crops"; crops.mkdir()
    review_dir = desktop_path() / f"{pdf_path.stem}_yutakaeng_review_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    log("PDF_OPEN", f"{pdf_path.name} を開いています")
    document = fitz.open(pdf_path)
    total = document.page_count
    all_rows: list[WireRow] = []
    # UIが総ページ数を早期に受け取り、初期推定残り時間を表示できるようにする。
    progress(0, total)
    order_no = "要確認"
    try:
        for page_index, page in enumerate(document, 1):
            log("PAGE_RENDER", f"ページ {page_index}/{total} を300dpiで解析しています")
            image = render_page(page)
            if page_index == 1:
                order_no = extract_order_number(image)
                log("ORDER_SCAN", f"オーダー番号候補: {order_no}")
            panel_no = extract_panel_number(image)
            log("PANEL_SCAN", f"盤番号候補: {panel_no or '未読取'}")
            # 主文字は画像層の単純切出しではなく、300dpi描画画像を用いるローカルONNX OCRで読む。
            # これにより画像層の反転・配置差で黒塗り相当を認識する不具合を回避する。
            frames = detect_frames(image)
            log("FRAME_DETECT", f"ページ {page_index}: {len(frames)}枠候補を検出")
            for index, frame in enumerate(frames, 1):
                all_rows.extend(analyze_frame(image, page_index, index, frame, crops, review_dir, panel_no, internal_only=internal_only, x_only=x_only))
            progress(page_index, total)
        if internal_only:
            log("RULE_FILTER", "盤内線モード: 端子台のI表記とIFブロック内の全マークを対象として出力")
        elif x_only:
            log("RULE_FILTER", "Xブロックモード: X1、X2などXブロック内の全マークを対象として出力")
        else:
            log("RULE_FILTER", "DT系端子台、D-線コード、端子台の盤内行きIを対象外として判定")
        log("SHEET_GROUP", "電線サイズ・コードごとにExcelシートを作成")
        output = write_excel(all_rows, pdf_path, order_no, desktop_path(), log, internal_only=internal_only, x_only=x_only)
        warning_count = sum(item.state == "警告あり" for item in all_rows)
        if warning_count and review_dir.exists():
            log("REVIEW_IMAGES_READY", f"警告セル画像 {warning_count}件を限定保存: {review_dir.name}")
        elif review_dir.exists():
            shutil.rmtree(review_dir, ignore_errors=True)
        elapsed = time.monotonic() - started
        log("PIPELINE_COMPLETE", f"完了: {len(all_rows)}行候補 / {elapsed:.1f}秒")
        return output
    finally:
        document.close()
        shutil.rmtree(work, ignore_errors=True)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="yutakaeng validation pipeline")
    parser.add_argument("pdf")
    parser.add_argument("--output", default=".")
    args = parser.parse_args()
    def _log(code: str, message: str) -> None: print(f"[{code}] {message}")
    path = run_pipeline(args.pdf, _log, lambda done, total: _log("PROGRESS", f"{done}/{total}"))
    print(path)
